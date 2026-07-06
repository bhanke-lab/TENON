# TENON - Translator Engine for Native Optimizer Notation
"""
translate.py - Stage 2: read a Baillie Run Set Up .xlsx, run the match
engine, and write a format-preserving copy with three columns appended
to the Lumber section: Comact Products | Instance IDs | Match Count.

Post-pass products (auto_activate + multi-destination union) aren't tied
to a single RSU row; they're appended to the FIRST Lumber row of the
matching thickness so the union of the sheet's lists equals match_all()
output for the fixture.

Usage:
    python translate.py --runsetup <file.xlsx> [--products <catalog.xml>]
                        [--mapping mapping.yaml] [--out <file.xlsx>]

Defaults: --products: newest tests/fixtures/_catalogs/allproducts_*.xml;
--out: <original>_comact.xlsx next to the input. Never overwrites input.
Known-zero inputs (9/4 rows; Beech/Butternut/Hickory runs) warn, not crash.

The Baillie template ships with sheet protection enabled; the output copy
would inherit it on round-trip, so protection is dropped on the run sheet
(other tabs, e.g. Data Validation, are left as-is).
"""
from __future__ import annotations

import argparse
import datetime as _dt
import sys
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT))

from src.parse_runsetup import load_runsetup_from_rows
from src.parse_products import load_products
from src.match import load_mapping, match_all

NEW_HEADERS = ("Comact Products", "Instance IDs", "Match Count")
# Excel column-width units per character at Aptos Narrow 16 (empirical;
# width units are calibrated to the default font, so big fonts need scaling).
CHAR_WIDTH_FACTOR = 1.5
# Template body font (Baillie Run Set Up uses Aptos Narrow 16 bold).
DATA_FONT = Font(name="Aptos Narrow", size=16, bold=True)
LINE_HEIGHT_PT = 21  # points per text line at 16pt
ZERO_FILL = PatternFill(fill_type="solid", start_color="FFFFC7CE",
                        end_color="FFFFC7CE")
ZERO_FONT = Font(name="Aptos Narrow", size=16, bold=True, color="FF9C0006")


def newest_catalog():
    catalogs = sorted(
        (ROOT / "tests" / "fixtures" / "_catalogs").glob("allproducts_*.xml")
    )
    return catalogs[-1] if catalogs else None


def _to_str(v):
    """Normalize an openpyxl cell value to the string form the CSV
    parser expects."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    if isinstance(v, bool):
        return "YES" if v else "NO"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, (_dt.datetime, _dt.date)):
        return f"{v.month}/{v.day}/{v.year}"
    return str(v).strip()


def sheet_grid(ws):
    return [[_to_str(c) for c in row] for row in ws.iter_rows(values_only=True)]


def find_lumber_header(grid):
    """Locate the Lumber header band by string match; absolute
    coordinates shift between templates."""
    for i, row in enumerate(grid):
        if row and row[0] == "Destination" and "Thickness" in row:
            return i
    return None


def pick_runsetup_sheet(wb):
    """The workbook carries extra tabs (Data_validation_on_runsetups).
    Pick the first sheet with a Lumber header band."""
    for ws in wb.worksheets:
        grid = sheet_grid(ws)
        if find_lumber_header(grid) is not None:
            return ws, grid
    return None, None


def main(argv=None):
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--runsetup", required=True, help="Run Set Up .xlsx")
    ap.add_argument("--products", default=None, help="AllProducts.xml")
    ap.add_argument("--mapping", default=str(ROOT / "mapping.yaml"))
    ap.add_argument("--out", default=None, help="Output .xlsx path")
    args = ap.parse_args(argv)

    runsetup_path = Path(args.runsetup)
    if not runsetup_path.exists():
        sys.exit(f"Run Set Up not found: {runsetup_path}")

    products_path = Path(args.products) if args.products else newest_catalog()
    if products_path is None or not products_path.exists():
        sys.exit(f"Catalog XML not found: {products_path}")

    out_path = (Path(args.out) if args.out
                else runsetup_path.with_name(runsetup_path.stem + "_comact.xlsx"))
    if out_path.resolve() == runsetup_path.resolve():
        sys.exit("Refusing to overwrite the input file.")

    wb = load_workbook(runsetup_path)
    ws, grid = pick_runsetup_sheet(wb)
    if ws is None:
        sys.exit("No sheet with a Lumber header band (Destination/Thickness) found.")

    # Drop inherited sheet protection so the output is editable.
    ws.protection.sheet = False

    rs = load_runsetup_from_rows(grid)
    if not rs.lumber_rows:
        sys.exit("Lumber header found but no Lumber rows parsed.")

    products = load_products(products_path)
    mapping = load_mapping(args.mapping)

    species_products = [p for p in products if p.species == rs.species]
    if not species_products:
        print(f"WARNING: no catalog products for species {rs.species!r} "
              f"({rs.species_raw}); every row will be 0-match. "
              "This species likely bypasses the Comact.")

    results, _w_unmapped, _l_unmapped, predicted = match_all(rs, products, mapping)

    # Bucket per worksheet row (grid index + 1 = ws row, 1-based).
    row_matches = {}
    matched_ids = set()
    for lrow, ms in results:
        row_matches[lrow.source_row + 1] = list(ms)
        matched_ids.update(p.instance_id for p in ms)

    # Post-pass products: attach to the first Lumber row of that thickness.
    unplaced = []
    for iid, p in sorted(predicted.items(), key=lambda kv: kv[1].name):
        if iid in matched_ids:
            continue
        target = next((lr for lr, _ in results if lr.thick == p.thick), None)
        if target is None:
            unplaced.append(p)
            continue
        row_matches[target.source_row + 1].append(p)
    for p in unplaced:
        print(f"WARNING: post-pass product has no Lumber row at its "
              f"thickness, not written to sheet: [{p.instance_id}] {p.name}")

    # Header band: append 3 styled columns to the right of the last
    # used Lumber header column.
    header_idx = find_lumber_header(grid)
    header_ws_row = header_idx + 1
    last_col = max(i for i, v in enumerate(grid[header_idx], start=1) if v)
    template = ws.cell(row=header_ws_row, column=last_col)
    for offset, title in enumerate(NEW_HEADERS, start=1):
        c = ws.cell(row=header_ws_row, column=last_col + offset, value=title)
        c.font = copy(template.font)
        c.fill = copy(template.fill)
        c.border = copy(template.border)
        c.alignment = copy(template.alignment)
    # Size the new columns off the longest content this run actually
    # produced, so no product name or ID ever wraps.
    placed = [p for plist in row_matches.values() for p in plist]
    name_w = max((len(p.name) for p in placed), default=20)
    id_w = max((len(p.instance_id) for p in placed), default=12)
    col_widths = (name_w * CHAR_WIDTH_FACTOR + 2,
                  id_w * CHAR_WIDTH_FACTOR + 2,
                  16)
    for offset, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(last_col + offset)].width = width

    # Data rows. Cants/Ties/Pallet rows are never touched: only rows the
    # Lumber parser emitted get written.
    zero_rows = []
    for lrow, _ in results:
        ws_row = lrow.source_row + 1
        plist = row_matches.get(ws_row, [])
        c1 = ws.cell(row=ws_row, column=last_col + 1,
                     value="\n".join(p.name for p in plist))
        c2 = ws.cell(row=ws_row, column=last_col + 2,
                     value="\n".join(p.instance_id for p in plist))
        c3 = ws.cell(row=ws_row, column=last_col + 3, value=len(plist))
        for c in (c1, c2, c3):
            c.font = copy(DATA_FONT)
        c1.alignment = Alignment(wrap_text=True, vertical="top")
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        c3.alignment = Alignment(vertical="top", horizontal="center")

        # Grow the row so every product line is visible. Columns are sized
        # so nothing wraps, but estimate rendered lines anyway as a belt
        # and suspenders. Never shrink a row the template made taller.
        chars_per_line = max(1, int((col_widths[0] - 2) / CHAR_WIDTH_FACTOR))
        lines = sum(max(1, -(-len(p.name) // chars_per_line)) for p in plist) or 1
        needed = lines * LINE_HEIGHT_PT + 6
        current = ws.row_dimensions[ws_row].height
        if current is None or needed > current:
            ws.row_dimensions[ws_row].height = needed

        if not plist:
            for c in (c1, c2, c3):
                c.fill = copy(ZERO_FILL)
            c3.font = copy(ZERO_FONT)
            zero_rows.append(lrow)

    for lrow in zero_rows:
        if lrow.thick == "9/4":
            reason = " (9/4 absent from Comact catalog: expected)"
        elif not species_products:
            reason = f" (no {rs.species} products in catalog: expected)"
        else:
            reason = " (rule gap or catalog gap: investigate)"
        print(f"WARNING: 0 matches on ws row {lrow.source_row + 1}: "
              f"{lrow.destination} {lrow.thick} {lrow.grade_code} "
              f"{lrow.color_code}{reason}")

    wb.save(out_path)
    print(f"\nWrote {out_path}")
    print(f"Run: {rs.species_raw} ({rs.species}), {rs.date}, "
          f"{len(results)} Lumber rows, "
          f"{len(predicted)} unique active products predicted.")


if __name__ == "__main__":
    main()
